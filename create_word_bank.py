"""
    Name: Devin DeLeon
    Date: 7/16/2023
    Description: This file will web scrape the given website and store 1000 words
    into a list. Then the program will store that list into an excel file named "test.xlsx"
"""

import os
import bs4
import requests
import openpyxl


def get_request_from_web():
    # getting our requests
    res = requests.get("https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/")
    res.raise_for_status()
    return res


def create_bs4_object(res):
    # creating a bs4 object for parsing
    bs4_object = bs4.BeautifulSoup(res.text, 'html.parser')
    return bs4_object


def create_word_list(bs4_object):
    word_list = bs4_object.select('p')[11].getText().split('\n')
    return word_list


def strip_word_list(word_list):
    word_list_stripped = []
    # strip the tab character at the beginning of each word
    for word in word_list:
        word_list_stripped.append(word.strip('\t'))
    return word_list_stripped


def does_word_bank_exist():
    # get the directory to check if xlsx file exists,  create new workbook if not, load if it does exist
    directory = os.listdir()
    if "Word_Bank.xlsx" in directory:
        print("word bank exists")
        return True
    else:
        print("no word bank found")
        return False


def create_word_bank(word_bank_bool, word_list_stripped):
    if word_bank_bool:
        wb = openpyxl.load_workbook("Word_Bank.xlsx")
    else:
        wb = openpyxl.Workbook()
    sheet_one = wb.active
    sheet_one.title = "Word Bank"
    for i in range(len(word_list_stripped)):
        sheet_one.cell(row=i + 1, column=1).value = word_list_stripped[i]
    print(sheet_one.title)
    wb.save("Word_Bank.xlsx")




