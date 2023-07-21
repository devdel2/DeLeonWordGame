"""
    Name: Devin DeLeon
    Date: 7/20/2023
    Description: This file will handle the logic for setting up the word game
"""

import openpyxl
import random


def print_title_statement():
    print("")
    print(" Welcome to the Word Guess Game! ".center(50, '-'))
    print("")


def print_game_rules():
    print("You will have unlimited attempts to guess the word.")
    print("Each guess is constituted by one letter")
    print("If the letter exists within the word, the letter will be revealed in the selected word.")
    print("If a letter doesn't exist within the word, the user will gain a strike.")
    print("After 5 strikes, the user will lose the game.")
    print("Guess the word before incurring 5 strikes and you win!\n")


def load_workbook():
    wb = openpyxl.load_workbook("Word_Bank.xlsx")
    sheet_one = wb.active
    return sheet_one


def select_random_word(wb_sheet):
    rand_index = random.randint(1, 1000)
    rand_word = wb_sheet.cell(row=rand_index, column=1).value
    return rand_word


def get_blank_spaces(rand_word):
    word_length = int(len(rand_word))
    blank_spaces = ''
    for i in range(1, word_length+1, 1):
        blank_spaces += '_'
    return blank_spaces

